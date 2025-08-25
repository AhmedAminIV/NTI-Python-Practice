"""
Exercise 1: You are asked to create a Python script that
cleans up old files from specific directories on the system.
The script should
1. Take a list of paths to monitor
2. Delete files older than a given number of days (e.g., 30 days).
3. Log all actions (what was deleted, when, from where).
4. Add an exclude list (files/folders to never delete)
"""
import os
import time
from pathlib import Path

# ------------------------------
# CONFIGURATION
# ------------------------------
paths_to_monitor = [
    "C:/temp",          # Example directory
    "C:/logs"           # Another directory
]

days_old = 30  # Delete files older than X days
exclude_list = [
    "important.txt",     # Exclude specific file
    "keep_folder"        # Exclude entire folder
]
log_file = "cleanup.log"


# ------------------------------
# FUNCTION DEFINITIONS
# ------------------------------
def is_excluded(path: Path):
    """Check if file or folder should be excluded"""
    for item in exclude_list:
        if path.name == item or item in path.parts:
            return True
    return False


def log_action(message: str):
    """Write a log entry"""
    with open(log_file, "a") as f:
        timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
        f.write(f"[{timestamp}] {message}\n")


def cleanup(path: Path):
    """Delete old files inside a path"""
    if not path.exists():
        log_action(f"Path not found: {path}")
        return

    cutoff_time = time.time() - (days_old * 86400)  # days → seconds

    for root, dirs, files in os.walk(path):
        root_path = Path(root)

        # Skip excluded directories
        if is_excluded(root_path):
            continue

        for file in files:
            file_path = root_path / file

            if is_excluded(file_path):
                continue

            # Check last modified time
            if file_path.stat().st_mtime < cutoff_time:
                try:
                    file_path.unlink()  # delete file
                    log_action(f"Deleted: {file_path}")
                except Exception as e:
                    log_action(f"Error deleting {file_path}: {e}")


# ------------------------------
# MAIN SCRIPT
# ------------------------------
if __name__ == "__main__":
    log_action("=== Cleanup Started ===")
    for path_str in paths_to_monitor:
        cleanup(Path(path_str))
    log_action("=== Cleanup Finished ===\n")

"""
Exercise 1.1: Log Rotation, Compression, and Archiving
1. Rotate logs → rename current log files with a timestamp suffix.
2. Compress old logs save space by zipping them (.gz or zip).
3. Archive logs move compressed logs to a separate folder (or S3).
4. Clean up archives older than a certain number of days.
"""
import time
import shutil
import gzip


BASE_PATH = Path(__file__).resolve().parent

# === Configuration ===
paths_to_monitor = [BASE_PATH / "logs"]
archive_path = BASE_PATH / "log_archive"
days_to_keep = 30                             # delete archives older than X days

# Ensure directories exist
for path in paths_to_monitor:
    path.mkdir(parents=True, exist_ok=True)
archive_path.mkdir(parents=True, exist_ok=True)


# === 1. Rotate Logs ===
def rotate_logs():
    for log_dir in paths_to_monitor:
        if log_dir.exists() and log_dir.is_dir():
            for log_file in log_dir.glob("*.log"):
                timestamp = time.strftime("%Y%m%d-%H%M%S")
                rotated_name = log_file.with_name(f"{log_file.stem}_{timestamp}{log_file.suffix}")
                log_file.rename(rotated_name)
                print(f"[ROTATED] {log_file} → {rotated_name}")
                compress_and_archive(rotated_name)
        else:
            print(f"[SKIPPED] {log_dir} does not exist or is not a directory.")


# === 2. Compress Old Logs ===
def compress_and_archive(rotated_file: Path):
    compressed_file = rotated_file.with_suffix(rotated_file.suffix + ".gz")

    with open(rotated_file, "rb") as f_in, gzip.open(compressed_file, "wb") as f_out:
        shutil.copyfileobj(f_in, f_out)

    rotated_file.unlink()  # remove original after compression
    print(f"[COMPRESSED] {rotated_file} → {compressed_file}")

    # === 3. Archive logs ===
    archived_file = archive_path / compressed_file.name
    compressed_file.rename(archived_file)
    print(f"[ARCHIVED] {archived_file}")


# === 4. Cleanup Old Archives ===
def cleanup_archives():
    cutoff_time = time.time() - (days_to_keep * 86400)

    for archive_file in archive_path.glob("*.gz"):
        if archive_file.stat().st_mtime < cutoff_time:
            try:
                archive_file.unlink()
                print(f"[DELETED] Old archive removed: {archive_file}")
            except Exception as e:
                print(f"[ERROR] Could not delete {archive_file}: {e}")


if __name__ == "__main__":
    print("=== Log Rotation and Archiving Started ===")
    rotate_logs()
    cleanup_archives()
    print("=== Log Rotation and Archiving Finished ===")


"""
Exercise 2:
Make a script to take all data from fake database and insert it into excel sheet
You are asked to write a Python script that:
1. Simulates a fake database (using SQLite).
2. Fetches data (like users, orders, or employees)
3. Writes the data into an Excel file.
4. Adds column headers, proper formatting, and maybe multiple sheets.
5. Logs the export process.
"""
import sqlite3
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from pathlib import Path
import logging

# === Paths ===
BASE_PATH = Path(__file__).resolve().parent
db_path = BASE_PATH / "fake_database.db"
excel_path = BASE_PATH / "exported_data.xlsx"
log_path = BASE_PATH / "export_log.txt"

# === Logging Setup ===
logging.basicConfig(
    filename=log_path,
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)


def create_fake_database():
    """Create fake SQLite database with sample data."""
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()

    # Create tables
    cur.execute("DROP TABLE IF EXISTS users;")
    cur.execute("DROP TABLE IF EXISTS orders;")

    cur.execute("""
        CREATE TABLE users (
            id INTEGER PRIMARY KEY,
            name TEXT,
            email TEXT
        )
    """)

    cur.execute("""
        CREATE TABLE orders (
            order_id INTEGER PRIMARY KEY,
            user_id INTEGER,
            product TEXT,
            amount REAL,
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    """)

    # Insert sample users
    cur.executemany("INSERT INTO users (name, email) VALUES (?, ?)", [
        ("Alice", "alice@example.com"),
        ("Bob", "bob@example.com"),
        ("Charlie", "charlie@example.com")
    ])

    # Insert sample orders
    cur.executemany("INSERT INTO orders (user_id, product, amount) VALUES (?, ?, ?)", [
        (1, "Laptop", 1200.50),
        (2, "Mouse", 25.99),
        (1, "Monitor", 300.00),
        (3, "Keyboard", 45.75),
    ])

    conn.commit()
    conn.close()
    logging.info("Fake database created with users and orders.")


def fetch_table_data(table_name):
    """Fetch all data from a given table."""
    conn = sqlite3.connect(db_path)
    cur = conn.cursor()
    cur.execute(f"SELECT * FROM {table_name}")
    rows = cur.fetchall()
    headers = [desc[0] for desc in cur.description]
    conn.close()
    logging.info(f"Fetched {len(rows)} rows from {table_name}.")
    return headers, rows


def write_to_excel(tables):
    """Write multiple tables into Excel with formatting."""
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    for table in tables:
        headers, rows = fetch_table_data(table)
        ws = wb.create_sheet(title=table.capitalize())

        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)

        # Write data rows
        for row_num, row_data in enumerate(rows, 2):
            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col, value=value)

        # Autofit column width
        for col in ws.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_length + 2

    wb.save(excel_path)
    logging.info(f"Data exported successfully to {excel_path}")


if __name__ == "__main__":
    logging.info("=== Export Process Started ===")
    create_fake_database()
    write_to_excel(["users", "orders"])
    logging.info("=== Export Process Finished ===")
    print(f"Export completed. Excel file created at: {excel_path}")
    print(f"Log file available at: {log_path}")


"""
Exercise 3:
Send System Storage Report via Email
You are asked to write a Python script that:
1. Checks system storage usage (free, used, total).
2. Checks memory and CPU usage (optional but useful).
3. Formats this data into a human-readable email (HTML or plain text).
4. Sends it via SMTP to a configured email address.
5. (Optional) Sends a warning only if disk usage exceeds a threshold (e.g., 80%).
6. (Optional) Attaches a CSV/Excel report for records.
"""
import psutil
import csv
from datetime import datetime
import os

# === Settings ===
THRESHOLD = 80  # % disk usage warning
REPORT_DIR = "reports"
os.makedirs(REPORT_DIR, exist_ok=True)  # create folder if not exists


# === Collect system stats ===
def get_system_stats():
    disk = psutil.disk_usage('/')
    memory = psutil.virtual_memory()
    cpu = psutil.cpu_percent(interval=1)

    stats = {
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "disk_total_gb": round(disk.total / (1024**3), 2),
        "disk_used_gb": round(disk.used / (1024**3), 2),
        "disk_free_gb": round(disk.free / (1024**3), 2),
        "disk_percent": disk.percent,
        "memory_total_gb": round(memory.total / (1024**3), 2),
        "memory_used_gb": round(memory.used / (1024**3), 2),
        "memory_percent": memory.percent,
        "cpu_percent": cpu,
    }
    return stats


# === Save to CSV ===
def save_report_csv(stats):
    filename = os.path.join(REPORT_DIR, "system_report.csv")
    file_exists = os.path.isfile(filename)

    with open(filename, "a", newline="") as csvfile:
        writer = csv.DictWriter(csvfile, fieldnames=stats.keys())
        if not file_exists:  # add header if new file
            writer.writeheader()
        writer.writerow(stats)

    return filename


# === Main ===
if __name__ == "__main__":
    stats = get_system_stats()

    # Print human-readable report
    print("\n=== System Storage Report ===")
    print(f"Timestamp: {stats['timestamp']}")
    print(f"Disk: {stats['disk_used_gb']}GB used / {stats['disk_total_gb']}GB "
          f"({stats['disk_percent']}%)")
    print(f"Memory: {stats['memory_used_gb']}GB used / {stats['memory_total_gb']}GB "
          f"({stats['memory_percent']}%)")
    print(f"CPU Usage: {stats['cpu_percent']}%")

    # Warning if above threshold
    if stats['disk_percent'] > THRESHOLD:
        print("⚠ WARNING: Disk usage exceeds threshold!")

    # Save report
    report_file = save_report_csv(stats)
    print(f"\nReport saved to: {report_file}")


"""
Exercise 4:
collect system logs every hour in a file with a name of date and hour file created.
make it elaborate.

Keyword: scheduled cron job (Linux, macOS), or Task Scheduler job (Windows)
"""
import os
import datetime
import psutil 
from pathlib import Path

# Define log directory
LOG_DIR = Path("logs")
LOG_DIR.mkdir(exist_ok=True)  # create folder if not exists

# Get current date and hour for filename
timestamp = datetime.datetime.now().strftime("%Y-%m-%d_%H")
log_file = LOG_DIR / f"system_log_{timestamp}.log"

# Collect system info
cpu_usage = psutil.cpu_percent(interval=1)
memory = psutil.virtual_memory()
disk = psutil.disk_usage('/')

# Prepare log content
log_content = f"""
===== System Log ({datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}) =====

CPU Usage: {cpu_usage}%
Memory: {memory.used / (1024**3):.2f} GB used / {memory.total / (1024**3):.2f} GB total
Disk: {disk.used / (1024**3):.2f} GB used / {disk.total / (1024**3):.2f} GB total

========================================================
"""

# Write to file (append mode if file exists)
with open(log_file, "a") as f:
    f.write(log_content)

print(f"Log written to {log_file}")


"""
Exercise 5: Detect Exposed Passwords or Secrets in Code. The task involves the following five steps:
1- Scanning code files (e.g., .py, .js, .env, .yml) for possible passwords, API keys, and tokens.
2- Using regex patterns to identify common exposures (e.g., password=..., api_key=..., secret=...).
3- Generating a detailed report that includes:
4- The file name, line number, and the suspicious content.
5- Supporting a whitelist/ignore list to reduce false positives.
"""

#!/usr/bin/env python3
import argparse, json, re, sys
from pathlib import Path
from typing import Iterable, List, Dict, Any

# ------------ Defaults ------------
DEFAULT_EXTS = {
    ".py", ".js", ".ts", ".tsx", ".jsx", ".env", ".yml", ".yaml", ".json",
    ".ini", ".cfg", ".txt", ".sh", ".ps1", ".xml", ".properties", ".toml"
}
DEFAULT_IGNORES = [
    ".git/**", "node_modules/**", "venv/**", ".venv/**", "dist/**", "build/**",
    "__pycache__/**", "*.min.js", "*.map", "*.lock", "*.zip", "*.gz", "*.tgz"
]

# Allowlist line-patterns (ignore if any matches)
ALLOWLIST_LINE_PATTERNS = [
    r"(?i)dummy|example|sample|fake|placeholder|test_value",
]

# Max file size to scan (bytes)
DEFAULT_MAX_BYTES = 2 * 1024 * 1024  # 2 MB

# ------------ Patterns (name -> regex) ------------
PATTERNS = {
    # Generic assignments
    "password_assignment": r"(?i)\b(pass(word|wd)?|pwd)\b\s*[:=]\s*['\"]?([^\s'\"#]{6,})",
    "api_key_assignment":  r"(?i)\b(api[_-]?key|secret|token|auth[_-]?token|access[_-]?token)\b\s*[:=]\s*['\"]?([A-Za-z0-9_\-]{8,})",
    # Cloud / services
    "aws_access_key_id":   r"\bAKIA[0-9A-Z]{16}\b",
    "aws_secret_key":      r"(?i)aws.{0,20}(secret|access).{0,5}[:=]\s*['\"]?[A-Za-z0-9/+=]{40}['\"]?",
    "google_api_key":      r"\bAIza[0-9A-Za-z\-_]{35}\b",
    "slack_token":         r"\bxox[baprs]-[0-9A-Za-z-]{10,48}\b",
    "jwt_token":           r"\beyJ[A-Za-z0-9_-]{10,}\.[A-Za-z0-9_-]{10,}\.[A-Za-z0-9_-]{10,}\b",
    # Private keys
    "private_key_block":   r"-----BEGIN (RSA|DSA|EC|OPENSSH|PGP) PRIVATE KEY-----",
    # Azure connection string (common shape)
    "azure_conn_string":   r"DefaultEndpointsProtocol=.*;AccountKey=[^;]+;.*",
}

COMPILED = {name: re.compile(rx) for name, rx in PATTERNS.items()}
ALLOWLIST_COMPILED = [re.compile(p) for p in ALLOWLIST_LINE_PATTERNS]

# ------------ Helpers ------------

def mask_secret(s: str, keep_start: int = 4, keep_end: int = 2) -> str:
    s = s.strip()
    if len(s) <= keep_start + keep_end:
        return "*" * len(s)
    return f"{s[:keep_start]}...{s[-keep_end:]}"


def load_ignore_globs(path: Path) -> List[str]:
    if not path or not path.exists():
        return []
    globs = []
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        globs.append(line)
    return globs


def is_ignored(path: Path, ignore_globs: Iterable[str]) -> bool:
    # Check against each glob from repo root
    for g in ignore_globs:
        if path.match(g) or any(p.match(g) for p in [path] + list(path.parents)):
            return True
    return False


def line_allowlisted(line: str) -> bool:
    return any(p.search(line) for p in ALLOWLIST_COMPILED)


def scan_file(path: Path, compiled: Dict[str, re.Pattern]) -> List[Dict[str, Any]]:
    findings = []
    try:
        # Skip binary-ish / too large
        if path.stat().st_size > DEFAULT_MAX_BYTES:
            return findings
        text = path.read_text(encoding="utf-8", errors="ignore")
    except Exception:
        return findings

    for lineno, line in enumerate(text.splitlines(), start=1):
        if line_allowlisted(line):
            continue
        for name, rx in compiled.items():
            for m in rx.finditer(line):
                # Capture secret content (prefer capturing group if present)
                groups = [g for g in m.groups() if g] or [m.group(0)]
                secret_snippet = groups[-1]  # likely the value
                findings.append({
                    "rule": name,
                    "file": str(path),
                    "line": lineno,
                    "match": mask_secret(secret_snippet),
                    "preview": line.strip()[:500],
                })
    return findings


def should_scan(path: Path, exts: Iterable[str]) -> bool:
    return path.suffix.lower() in exts

# ------------ Main ------------
def main():
    ap = argparse.ArgumentParser(
        description="Detect exposed passwords/API keys/tokens in source trees."
    )
    ap.add_argument("--root", default=".", help="Root directory to scan (default: .)")
    ap.add_argument("--exts", default=",".join(sorted(DEFAULT_EXTS)),
                    help="Comma-separated list of file extensions to scan.")
    ap.add_argument("--ignore-file", default=".secretignore",
                    help="Path to an ignore file with glob patterns (default: .secretignore).")
    ap.add_argument("--format", choices=["csv", "json"], default="csv",
                    help="Report format (csv or json).")
    ap.add_argument("--output", default="secret_report.csv",
                    help="Output file name (default: secret_report.csv / .json).")
    ap.add_argument("--fail-on-find", action="store_true",
                    help="Exit with code 1 if any findings are detected.")
    ap.add_argument("--verbose", action="store_true", help="Print matches to stdout.")
    args = ap.parse_args()

    root = Path(args.root).resolve()
    exts = {e if e.startswith(".") else f".{e}" for e in args.exts.split(",") if e.strip()}
    ignore_globs = DEFAULT_IGNORES + load_ignore_globs((root / args.ignore_file))

    all_findings: List[Dict[str, Any]] = []

    for p in root.rglob("*"):
        if not p.is_file():
            continue
        if is_ignored(p.relative_to(root), ignore_globs):
            continue
        if not should_scan(p, exts):
            continue
        file_findings = scan_file(p, COMPILED)
        all_findings.extend(file_findings)
        if args.verbose:
            for f in file_findings:
                print(f"[{f['rule']}] {f['file']}:{f['line']} :: {f['match']}")

    # Write report
    out = Path(args.output)
    if args.format == "json":
        if out.suffix.lower() != ".json":
            out = out.with_suffix(".json")
        out.write_text(json.dumps(all_findings, indent=2), encoding="utf-8")
    else:
        # CSV
        if out.suffix.lower() != ".csv":
            out = out.with_suffix(".csv")
        import csv
        with out.open("w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=["rule", "file", "line", "match", "preview"])
            writer.writeheader()
            writer.writerows(all_findings)

    # Summary
    print(f"\nScanned root: {root}")
    print(f"Extensions: {', '.join(sorted(exts))}")
    print(f"Ignores: {len(ignore_globs)} patterns")
    print(f"Findings: {len(all_findings)}")
    print(f"Report: {out}")

    if args.fail_on_find and all_findings:
        sys.exit(1)

if __name__ == "__main__":
    main()
